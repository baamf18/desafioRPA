import pandas as pd
import xml.etree.ElementTree as ET
import os

import openpyxl

class Read_xml():
    def __init__(self, directory) -> None:
        self.directory = directory

    def all_files(self):
        return [os.path.join(self.directory, arq) for arq in os.listdir(self.directory)
            if arq.lower().endswith(".xml")]

    def nfe_data(self, xml):
        root = ET.parse(xml).getroot()

        nsNFE = {'ns': "http://www.portalfiscal.inf.br/nfe"}

        #Fornecedor
        fornecedor = self.check_none(root.find('ns:NFe/ns:infNFe/ns:emit/ns:xNome', nsNFE))

        #Data Emissão
        dataEmissao_nfe = self.check_none(root.find('ns:NFe/ns:infNFe/ns:ide/ns:dhEmi', nsNFE))

        #CFOP
        CFOP_nfe = self.check_none(root.find('ns:NFe/ns:infNFe/ns:det/ns:prod/ns:CFOP', nsNFE))

        #status
        status_nfe = self.check_none(root.find('ns:protNFe/ns:infProt/ns:cStat', nsNFE))

        dados = [{'fornecedor': fornecedor, 'data emissao': dataEmissao_nfe, 'CFOP': CFOP_nfe, 'status': status_nfe}]

        return dados

    def check_none(self, var):
        if var == None:
            return ""
        else:
            return var.text


if __name__ == "__main__":
    xml = Read_xml(r'C:\Users\cnmfe\Downloads\Finance_RPA_Developer-20220801T203231Z-001\Finance_RPA_Developer\NFes-35755699000184_(1)\Autorizada\Recebidas')
    all = xml.all_files()

    listas = []

    for i in all:
        result = xml.nfe_data(i)
        listas += result
        
    df = pd.DataFrame(listas)

with pd.ExcelWriter('Omie_Contas_Pagar_v1_1_1.xlsx', mode = 'a', engine = 'openpyxl', if_sheet_exists = 'overlay') as writer:
    # Pedreiragem master forçando a linha de início para escrever no excel ser a linha 5, pois o mode append não está funcionando:
      df['fornecedor'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 2)
      df['data emissao'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 8)
      df['CFOP'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 20)
      df['status'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 48)

if __name__ == "__main__":
    xml = Read_xml(r'C:\Users\cnmfe\Downloads\Finance_RPA_Developer-20220801T203231Z-001\Finance_RPA_Developer\NFes-35755699000184_(2)\Autorizada\Recebidas')
    all = xml.all_files()

    listas = []

    for i in all:
        result = xml.nfe_data(i)
        listas += result
        
    df = pd.DataFrame(listas)

with pd.ExcelWriter('Omie_Contas_Pagar_v1_1_1.xlsx', mode = 'a', engine = 'openpyxl', if_sheet_exists = 'overlay') as writer:
    # Pedreiragem master forçando a linha de início para escrever no excel ser a linha 5, pois o mode append não está funcionando:
      df['fornecedor'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 2)
      df['data emissao'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 8)
      df['CFOP'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 20)
      df['status'].to_excel(writer, index = False, header = False, sheet_name = 'Omie_Contas_Pagar', startrow = writer.sheets['Omie_Contas_Pagar'].max_row, startcol = 48)

